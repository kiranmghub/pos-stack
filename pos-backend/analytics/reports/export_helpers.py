# analytics/reports/export_helpers.py
"""
Export helper functions for report generation.
Provides utilities for converting report data to CSV, Excel, and PDF formats.
"""
import logging
import csv
import io
from typing import Dict, List, Any, Optional
from decimal import Decimal
from datetime import datetime

logger = logging.getLogger(__name__)


def export_report_to_csv(report_data: Dict[str, Any], report_type: str) -> str:
    """
    Export report data to CSV format.
    
    Args:
        report_data: Report data dictionary
        report_type: Type of report (e.g., "sales", "products", "financial")
    
    Returns:
        CSV string
    """
    output = io.StringIO()
    
    if report_type == "sales":
        # Export sales detail data
        if "results" in report_data:
            rows = report_data["results"]
            if rows:
                fieldnames = list(rows[0].keys())
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for row in rows:
                    csv_row = {}
                    for key, value in row.items():
                        if isinstance(value, (Decimal, float)):
                            csv_row[key] = str(value)
                        elif isinstance(value, datetime):
                            csv_row[key] = value.isoformat()
                        elif value is None:
                            csv_row[key] = ""
                        else:
                            csv_row[key] = str(value)
                    writer.writerow(csv_row)
    
    elif report_type == "products":
        # Export top products
        if "top_products_by_revenue" in report_data:
            rows = report_data["top_products_by_revenue"]
            if rows:
                writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                for row in rows:
                    csv_row = {k: str(v) if v is not None else "" for k, v in row.items()}
                    writer.writerow(csv_row)
    
    elif report_type == "financial":
        # Export financial summary - multiple sections
        sections = []
        if "summary" in report_data:
            sections.append(("Summary", [report_data["summary"]]))
        if "payment_methods" in report_data:
            sections.append(("Payment Methods", report_data["payment_methods"]))
        if "discount_rules" in report_data:
            sections.append(("Discount Rules", report_data["discount_rules"]))
        if "tax_rules" in report_data:
            sections.append(("Tax Rules", report_data["tax_rules"]))
        
        for section_name, rows in sections:
            if rows:
                output.write(f"\n{section_name}\n")
                output.write("=" * 50 + "\n")
                fieldnames = list(rows[0].keys())
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for row in rows:
                    csv_row = {k: str(v) if v is not None else "" for k, v in row.items()}
                    writer.writerow(csv_row)
                output.write("\n")
    
    elif report_type == "customers":
        # Export top customers
        if "top_customers" in report_data:
            rows = report_data["top_customers"]
            if rows:
                writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                for row in rows:
                    csv_row = {k: str(v) if v is not None else "" for k, v in row.items()}
                    writer.writerow(csv_row)
    
    elif report_type == "employees":
        # Export top employees
        if "top_employees" in report_data:
            rows = report_data["top_employees"]
            if rows:
                writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
                writer.writeheader()
                for row in rows:
                    csv_row = {k: str(v) if v is not None else "" for k, v in row.items()}
                    writer.writerow(csv_row)
    
    elif report_type == "returns":
        # Export returns data
        sections = []
        if "summary" in report_data:
            sections.append(("Summary", [report_data["summary"]]))
        if "reason_breakdown" in report_data:
            sections.append(("Reason Breakdown", report_data["reason_breakdown"]))
        if "disposition_breakdown" in report_data:
            sections.append(("Disposition Breakdown", report_data["disposition_breakdown"]))
        if "status_breakdown" in report_data:
            sections.append(("Status Breakdown", report_data["status_breakdown"]))
        
        for section_name, rows in sections:
            if rows:
                output.write(f"\n{section_name}\n")
                output.write("=" * 50 + "\n")
                fieldnames = list(rows[0].keys())
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                for row in rows:
                    csv_row = {k: str(v) if v is not None else "" for k, v in row.items()}
                    writer.writerow(csv_row)
                output.write("\n")
    
    return output.getvalue()


def export_report_to_excel(report_data: Dict[str, Any], report_type: str, tenant_name: str) -> bytes:
    """
    Export report data to Excel format using openpyxl.
    
    Args:
        report_data: Report data dictionary
        report_type: Type of report
        tenant_name: Tenant name for header
    
    Returns:
        Excel file as bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Please install it.")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    def create_sheet(name: str, rows: List[Dict[str, Any]], title: Optional[str] = None):
        """Create a worksheet with data."""
        if not rows:
            return
        
        ws = wb.create_sheet(title=name)
        
        # Add title if provided
        if title:
            ws.append([title])
            ws.merge_cells(f"A1:{get_column_letter(len(rows[0].keys()))}1")
            title_cell = ws["A1"]
            title_cell.font = Font(bold=True, size=14)
            ws.append([])  # Empty row
        
        # Add headers
        fieldnames = list(rows[0].keys())
        ws.append(fieldnames)
        header_row = ws[ws.max_row]
        for cell in header_row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for row in rows:
            values = []
            for field in fieldnames:
                value = row.get(field, "")
                if isinstance(value, (Decimal, float)):
                    values.append(float(value))
                elif isinstance(value, datetime):
                    values.append(value.isoformat())
                elif value is None:
                    values.append("")
                else:
                    values.append(str(value))
            ws.append(values)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate sheets based on report type
    if report_type == "sales" and "results" in report_data:
        create_sheet("Sales Detail", report_data["results"], f"Sales Report - {tenant_name}")
    
    elif report_type == "products" and "top_products_by_revenue" in report_data:
        create_sheet("Top Products", report_data["top_products_by_revenue"], f"Product Report - {tenant_name}")
    
    elif report_type == "financial":
        if "summary" in report_data:
            create_sheet("Summary", [report_data["summary"]], f"Financial Report - {tenant_name}")
        if "payment_methods" in report_data and report_data["payment_methods"]:
            create_sheet("Payment Methods", report_data["payment_methods"])
        if "discount_rules" in report_data and report_data["discount_rules"]:
            create_sheet("Discount Rules", report_data["discount_rules"])
        if "tax_rules" in report_data and report_data["tax_rules"]:
            create_sheet("Tax Rules", report_data["tax_rules"])
    
    elif report_type == "customers" and "top_customers" in report_data:
        create_sheet("Top Customers", report_data["top_customers"], f"Customer Report - {tenant_name}")
    
    elif report_type == "employees" and "top_employees" in report_data:
        create_sheet("Top Employees", report_data["top_employees"], f"Employee Report - {tenant_name}")
    
    elif report_type == "returns":
        if "summary" in report_data:
            create_sheet("Summary", [report_data["summary"]], f"Returns Report - {tenant_name}")
        if "reason_breakdown" in report_data and report_data["reason_breakdown"]:
            create_sheet("By Reason", report_data["reason_breakdown"])
        if "disposition_breakdown" in report_data and report_data["disposition_breakdown"]:
            create_sheet("By Disposition", report_data["disposition_breakdown"])
        if "status_breakdown" in report_data and report_data["status_breakdown"]:
            create_sheet("By Status", report_data["status_breakdown"])
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_report_to_pdf(report_data: Dict[str, Any], report_type: str, tenant_name: str, date_range: str) -> bytes:
    """
    Export report data to PDF format using reportlab.
    
    Args:
        report_data: Report data dictionary
        report_type: Type of report
        tenant_name: Tenant name for header
        date_range: Date range string for header
    
    Returns:
        PDF file as bytes
    """
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        raise ImportError("reportlab is required for PDF export. Please install it.")
    
    # Handle macOS/hashlib.md5 issue
    try:
        import hashlib as _hashlib
        _hashlib.md5(usedforsecurity=False)
    except TypeError:
        from reportlab.pdfbase import pdfdoc as _pdfdoc
        def _md5_no_kw(*args, **kwargs):
            return _hashlib.md5(*args)
        _pdfdoc.md5 = _md5_no_kw
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=18,
        textColor=colors.HexColor("#1E293B"),
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=colors.HexColor("#334155"),
        spaceAfter=8,
    )
    normal_style = styles["Normal"]
    
    story = []
    
    # Title
    story.append(Paragraph(f"<b>{report_type.title()} Report</b>", title_style))
    story.append(Paragraph(f"<b>Tenant:</b> {tenant_name}", normal_style))
    story.append(Paragraph(f"<b>Date Range:</b> {date_range}", normal_style))
    story.append(Spacer(1, 0.2 * inch))
    
    def create_table(data: List[Dict[str, Any]], title: Optional[str] = None):
        """Create a PDF table from data."""
        if not data:
            return
        
        if title:
            story.append(Paragraph(f"<b>{title}</b>", heading_style))
            story.append(Spacer(1, 0.1 * inch))
        
        # Prepare table data
        fieldnames = list(data[0].keys())
        table_data = [fieldnames]  # Header row
        
        for row in data:
            table_data.append([
                str(row.get(field, "")) if row.get(field) is not None else ""
                for field in fieldnames
            ])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            # Data rows
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))
    
    # Generate content based on report type
    if report_type == "sales" and "results" in report_data:
        create_table(report_data["results"], "Sales Detail")
    
    elif report_type == "products" and "top_products_by_revenue" in report_data:
        create_table(report_data["top_products_by_revenue"], "Top Products by Revenue")
    
    elif report_type == "financial":
        if "summary" in report_data:
            create_table([report_data["summary"]], "Summary")
        if "payment_methods" in report_data and report_data["payment_methods"]:
            create_table(report_data["payment_methods"], "Payment Methods")
        if "discount_rules" in report_data and report_data["discount_rules"]:
            create_table(report_data["discount_rules"], "Discount Rules")
        if "tax_rules" in report_data and report_data["tax_rules"]:
            create_table(report_data["tax_rules"], "Tax Rules")
    
    elif report_type == "customers" and "top_customers" in report_data:
        create_table(report_data["top_customers"], "Top Customers")
    
    elif report_type == "employees" and "top_employees" in report_data:
        create_table(report_data["top_employees"], "Top Employees")
    
    elif report_type == "returns":
        if "summary" in report_data:
            create_table([report_data["summary"]], "Summary")
        if "reason_breakdown" in report_data and report_data["reason_breakdown"]:
            create_table(report_data["reason_breakdown"], "Breakdown by Reason")
        if "disposition_breakdown" in report_data and report_data["disposition_breakdown"]:
            create_table(report_data["disposition_breakdown"], "Breakdown by Disposition")
        if "status_breakdown" in report_data and report_data["status_breakdown"]:
            create_table(report_data["status_breakdown"], "Breakdown by Status")
    
    # Build PDF
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

